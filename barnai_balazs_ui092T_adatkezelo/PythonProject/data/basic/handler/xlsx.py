import os
from openpyxl import Workbook
from basic import generator
from basic.model_dataclasses import Person, Bicycle, Laptop

#  KAPCSOLATOK
def link_people_with_items(people: list[Person],
                           bicycles: list[Bicycle],
                           laptops: list[Laptop]) -> None:
    """Beállítja a kapcsolatokat a példányok között."""
    people_by_id = {p.id: p for p in people}

    # Biciklik hozzárendelése tulajdonoshoz
    for b in bicycles:
        owner = people_by_id.get(b.owner_id)
        if owner:
            b.owner = owner
            owner.bicycles.append(b)

    # Laptopok hozzárendelése tulajdonoshoz
    for l in laptops:
        owner = people_by_id.get(l.owner_id)
        if owner:
            l.owner = owner
            owner.laptops.append(l)


#  XLSX Írás
def write_people(people, workbook, sheet_name="people", heading=True):
    sheet_name = sheet_name or "people"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    if heading:
        sheet.append(["id", "name", "age", "male"])
    for p in people:
        sheet.append([p.id, p.name, p.age, p.male])


def write_bicycles(bicycles, workbook, sheet_name="bicycles", heading=True):
    sheet_name = sheet_name or "bicycles"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    if heading:
        sheet.append(["id", "brand", "model", "year", "owner_id"])
    for b in bicycles:
        owner_id = b.owner.id if hasattr(b, "owner") and b.owner else b.owner_id
        sheet.append([b.id, b.brand, b.model, b.year, owner_id])


def write_laptops(laptops, workbook, sheet_name="laptops", heading=True):
    sheet_name = sheet_name or "laptops"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    if heading:
        sheet.append(["id", "brand", "model", "year", "ram", "vram", "owner_id"])
    for l in laptops:
        owner_id = l.owner.id if hasattr(l, "owner") and l.owner else l.owner_id
        sheet.append([l.id, l.brand, l.model, l.year, l.ram, l.vram, owner_id])


def write_relations(people, workbook, sheet_name="relations"):
    """Kapcsolatok kimentése egy külön sheet-be."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["person_id", "person_name", "bicycle_models", "laptop_models"])

    for p in people:
        bicycles = ", ".join([b.model for b in p.bicycles]) if hasattr(p, "bicycles") else ""
        laptops = ", ".join([l.model for l in p.laptops]) if hasattr(p, "laptops") else ""
        sheet.append([p.id, p.name, bicycles, laptops])


# Közvetítő Függvény
def write_entities(entities, workbook, sheet_name=None):
    if not entities:
        return
    first = entities[0]
    if isinstance(first, Person):
        write_people(entities, workbook, sheet_name)
    elif isinstance(first, Bicycle):
        write_bicycles(entities, workbook, sheet_name)
    elif isinstance(first, Laptop):
        write_laptops(entities, workbook, sheet_name)
    else:
        raise TypeError(f"Ismeretlen típus: {type(first)}")


#  MAIN
if __name__ == "__main__":
    save_path = "C:/hallgato"
    os.makedirs(save_path, exist_ok=True)

    wb = Workbook()
    # Alapértelmezett "Sheet" törlése
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # adat generalas
    people = generator.generate_people(80)
    bicycles = generator.generate_bicycles(50, people)
    laptops = generator.generate_laptops(50, people)

    # Kapcsolatok beállítása (1:N)
    link_people_with_items(people, bicycles, laptops)

    # Írás a munkafüzetbe
    write_entities(people, wb)
    write_entities(bicycles, wb)
    write_entities(laptops, wb)
    write_relations(people, wb)

    # Fájl mentése
    excel_file = os.path.join(save_path, "data_xlsx.xlsx")
    wb.save(excel_file)
    print(f"Excel fájl létrehozva: {excel_file}")

    # Ellenőrzés
    print("\n=== KAPCSOLATOK ===")
    for p in people:
        print(f"{p.name}:")
        print("  Biciklik:", [b.model for b in p.bicycles])
        print("  Laptopok:", [l.model for l in p.laptops])
